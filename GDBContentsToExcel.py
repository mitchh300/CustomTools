""" Writes geodatabase contents to Excel (.xlsx)
    All feature classes, tables, relationship classes
    will have Describe object properties written to respective
    Excel columns.  Also, field names and types
    will be written after each dataset info is written.

    Author: Mitch Holley
    Date: 10/25/2017
    Last Edits: 8/30/2018
    Python Version 2.7.10
"""
import arcpy
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime

#date string
today = datetime.datetime.now().strftime('%m%d%Y')

#input geodatabase
InputGDB = r'path\to\your\geodatabase.gdb'

#env
arcpy.env.workspace = InputGDB

#find all datasets
FeatureSets = [x for x in arcpy.ListDatasets("","Feature")]

#find all relationship classes in FeatureDatasets
RelSets = []
for x in FeatureSets:
    desc = arcpy.Describe(x)
    for item in desc.children:
        if item.datasetType == 'RelationshipClass':
            RelSets.append((item,desc.name))
        
#excel items
wb = Workbook()
ws1 = wb.active
ws1.title = 'FeatureClasses'
ws2 = wb.create_sheet(title = 'Tables')
ws3 = wb.create_sheet(title = 'RelationshipClasses')

ws1.freeze_panes = 'A2'
ws2.freeze_panes = 'A2'
ws3.freeze_panes = 'A2'

underline = Font(underline='single')
bold = Font(bold = True)

#FeatureClasses schema
ws1['A1'] = 'Name'
ws1['B1'] = 'Type'
ws1['C1'] = 'HasZ'
ws1['D1'] = 'HasM'
ws1['E1'] = 'CoordinateSystem'
ws1['F1'] = 'FeatureDataset'
ws1['G1'] = 'EditorTracking'
ws1['H1'] = 'CatalogPath'

ws1['A1'].font = underline
ws1['B1'].font = underline
ws1['C1'].font = underline
ws1['D1'].font = underline
ws1['E1'].font = underline
ws1['F1'].font = underline
ws1['G1'].font = underline
ws1['H1'].font = underline

#Tables schema
ws2['A1'] = 'Name'
ws2['B1'] = 'Alias'
ws2['C1'] = 'EditorTracking'
ws2['D1'] = 'CatalogPath'

ws2['A1'].font = underline
ws2['B1'].font = underline
ws2['C1'].font = underline
ws2['D1'].font = underline

#RelationshipClasses schema
ws3['A1'] = 'Name'
ws3['B1'] = 'FeatureDataset'
ws3['C1'] = 'Cardinality'
ws3['D1'] = 'OriginFeatureClass'
ws3['E1'] = 'PrimaryKey'
ws3['F1'] = 'DestinationFeatureClass'
ws3['G1'] = 'ForeignKey'
ws3['H1'] = 'Rules'

ws2['A1'].font = underline
ws3['B1'].font = underline
ws3['C1'].font = underline
ws3['D1'].font = underline
ws3['E1'].font = underline
ws3['F1'].font = underline
ws3['G1'].font = underline
ws3['H1'].font = underline

#list FC in datasets, assign the return values to ws1 rows
idx = 2
if FeatureSets:
    for ds in FeatureSets:
        for fc in sorted(arcpy.ListFeatureClasses("","",ds)):
            desc = arcpy.Describe(fc)
            ws1['A{0}'.format(idx)] = fc
            ws1['A{0}'.format(idx)].font = bold
            ws1['B{0}'.format(idx)] = desc.shapeType
            ws1['C{0}'.format(idx)] = 'Enabled' if desc.hasZ else 'Disabled'
            ws1['D{0}'.format(idx)] = 'Enabled' if desc.hasM else 'Disabled'
            ws1['E{0}'.format(idx)] = desc.spatialReference.name
            ws1['F{0}'.format(idx)] = ds
            ws1['G{0}'.format(idx)] = 'Enabled' if desc.editorTrackingEnabled else 'Disabled'
            ws1['H{0}'.format(idx)] = desc.catalogPath
            idx += 1
            fields = [x for x in arcpy.ListFields(fc)]
            for field in fields:
                ws1['A{}'.format(idx)] = field.name
                ws1['B{}'.format(idx)] = field.type
                idx += 1
            idx += 1
            
#list FC not in dataset, assign the return values to ws1 rows
for fc in sorted(arcpy.ListFeatureClasses()):
    desc = arcpy.Describe(fc)
    ws1['A{0}'.format(idx)] = fc
    ws1['A{0}'.format(idx)].font = bold
    ws1['B{0}'.format(idx)] = desc.shapeType
    ws1['C{0}'.format(idx)] = 'Enabled' if desc.hasZ else 'Disabled'
    ws1['D{0}'.format(idx)] = 'Enabled' if desc.hasM else 'Disabled'
    ws1['E{0}'.format(idx)] = desc.spatialReference.name
    ws1['F{0}'.format(idx)] = '-'
    ws1['G{0}'.format(idx)] = 'Enabled' if desc.editorTrackingEnabled else 'Disabled'
    ws1['H{0}'.format(idx)] = desc.catalogPath
    idx += 1
    fields = [x for x in arcpy.ListFields(fc)]
    for field in fields:
        ws1['A{}'.format(idx)] = field.name
        ws1['B{}'.format(idx)] = field.type
        idx += 1
    idx += 1

#list tables in gdb, assign the return values to ws2 rows
ws2 = wb['Tables']
idx = 2
for table in arcpy.ListTables():
    desc = arcpy.Describe(table)
    ws2['A{}'.format(idx)] = desc.name
    ws2['A{}'.format(idx)].font = bold
    ws2['B{}'.format(idx)] = desc.aliasName if desc.aliasName else '-'
    ws2['C{}'.format(idx)] = 'Enabled' if desc.editorTrackingEnabled else 'Disabled'
    ws2['D{}'.format(idx)] = desc.catalogPath
    idx += 1
    fields = [x for x in arcpy.ListFields(table)]
    for field in fields:
        ws2['A{}'.format(idx)] = field.name
        ws2['B{}'.format(idx)] = field.type
        idx += 1
    idx += 1

if RelSets:
    idx = 2
    ws3 = wb['RelationshipClasses']
    for rel in RelSets:
        obj = rel[0]
        ws3['A{}'.format(idx)] = obj.name
        ws3['B{}'.format(idx)] = rel[1]
        ws3['C{}'.format(idx)] = obj.cardinality
        ws3['D{}'.format(idx)] = ", ".join(x for x in obj.originClassNames)
        ws3['E{}'.format(idx)] = ''.format(x[0] for x in obj.originClassKeys[0])
        ws3['F{}'.format(idx)] = ", ".join(obj.destinationClassNames)
        ws3['G{}'.format(idx)] = [x[0] for x in obj.originClassKeys[1]]
        ws3['H{}'.format(idx)] = obj.relationshipRules
        idx += 1

#save output, delete previous version if found
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
outputXLS = os.path.join(desktop, 'GDB_Contents_{}.xlsx'.format(today))
if os.path.exists(outputXLS):
    os.remove(outputXLS)
wb.save(outputXLS)
print'Complete'
